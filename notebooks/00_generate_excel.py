# Databricks notebook source
# MAGIC %pip install openpyxl google-cloud-storage

# COMMAND ----------
# MAGIC %md
# MAGIC # 📥 Stage 0: Ingestion Mock Data Generator
# MAGIC 
# MAGIC Generates mock sales Excel files for countries and distributors, and uploads them to the GCS `landing/` directory.

# COMMAND ----------
import io
import os
import json
import random
import logging
from datetime import date
import openpyxl
from openpyxl.styles import Font
from google.cloud import storage
from google.oauth2 import service_account

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("excel_generator")

# COMMAND ----------
# Define widgets
dbutils.widgets.text("env", "dv")
dbutils.widgets.text("source_bucket", "")
dbutils.widgets.text("run_date", "")

env = dbutils.widgets.get("env")
source_bucket_name = dbutils.widgets.get("source_bucket")
run_date = dbutils.widgets.get("run_date")

if not source_bucket_name:
    raise ValueError("Parameter 'source_bucket' is required.")

today_yyyymmdd = run_date.replace("-", "") if run_date else date.today().strftime("%Y%m%d")
logger.info(f"Mock Excel generation started for Env: {env} | Bucket: {source_bucket_name} | Date: {today_yyyymmdd}")

# COMMAND ----------
# Configure GCS Authentication using Service Account Key
# PASTE YOUR GCP SERVICE ACCOUNT JSON KEY CONTENT INSIDE THE TRIPLE QUOTES BELOW
gcp_key_json = """"""

try:
    if gcp_key_json.strip():
        info = json.loads(gcp_key_json)
        credentials = service_account.Credentials.from_service_account_info(info)
        client = storage.Client(credentials=credentials, project=info.get("project_id"))
        logger.info("Successfully authenticated to GCS storage client using direct JSON key.")
    else:
        logger.warning("gcp_key_json variable is empty. Attempting default credentials.")
        client = storage.Client()
except Exception as e:
    logger.error(f"Authentication failed: {e}. Falling back to default credentials.")
    client = storage.Client()

bucket = client.bucket(source_bucket_name)

# COMMAND ----------
# Catalog and constants
CATALOGUE_INR = {
    "Samsung": [
        ("Galaxy A56 5G", 38990, "1029384756102938"),
        ("Galaxy M17 5G", 13965, "1928374650192837"),
        ("Galaxy S26 Ultra", 121870, "1938475620193847"),
        ("Galaxy F07", 9499, "2039485710293847"),
        ("Galaxy S26", 76500, "2948571029384756"),
        ("Galaxy F34 5G", 21999, "3847562910384756"),
        ("Galaxy S24 Ultra 5G", 134999, "4729104857291038"),
        ("Galaxy A36 5G", 35208, "4857291038475629"),
        ("Galaxy M35 5G", 17499, "5729103847562910"),
        ("Galaxy S22 5G", 85999, "5829103847562910"),
        ("Galaxy M56 5G", 23499, "7584930218475629"),
        ("Galaxy Z Fold 5", 159999, "8392048175930284"),
        ("Galaxy A07", 10300, "8475629103847562"),
        ("Galaxy S24 FE", 39999, "9283746501928374"),
        ("Galaxy A06", 7999, "9384756201938475"),
    ],
    "Vivo": [
        ("X Fold3 Pro", 159999, "3829104857291030"),
        ("X100 Pro", 89999, "5729104857291031"),
        ("X80 Pro", 79999, "1938475620193841"),
        ("X90 Pro", 74999, "5829103847562911"),
        ("X100", 63999, "2948571029384751"),
        ("V30 Pro", 41999, "9283746501928371"),
        ("V30", 33999, "1029384756102931"),
        ("V29 5G", 32999, "4857291038475621"),
        ("T2 Pro 5G", 23999, "7584930218475621"),
        ("Y200 5G", 21999, "3847562910384751"),
        ("Y56 5G", 16999, "5729103847562911"),
        ("Y28 5G", 13999, "1928374650192831"),
        ("T2x 5G", 12999, "8475629103847561"),
        ("Y16", 10499, "2039485710293841"),
        ("Y02t", 8999, "9384756201938471"),
    ],
    "Oppo": [
        ("Find N3 Fold", 149999, "8392048175930281"),
        ("Find X7 Ultra", 99999, "4729104857291031"),
        ("Find N2 Flip", 89999, "1938475620193841"),
        ("Find X6 Pro", 84999, "5829103847562911"),
        ("Reno 10 Pro+ 5G", 54999, "2948571029384751"),
        ("Reno 11 Pro 5G", 39999, "9283746501928371"),
        ("Reno 10 5G", 32999, "1029384756102931"),
        ("Reno 11 5G", 29999, "4857291038475621"),
        ("F23 5G", 24999, "7584930218475621"),
        ("F25 Pro 5G", 23999, "3847562910384751"),
        ("A79 5G", 19999, "5729103847562911"),
        ("A59 5G", 14999, "1928374650192831"),
        ("A58", 13999, "8475629103847561"),
        ("A38", 12999, "2039485710293841"),
        ("A18", 9999, "9384756201938471"),
    ],
    "OnePlus": [
        ("OnePlus Open", 139999, "8392048175930280"),
        ("OnePlus 10 Pro", 66999, "4729104857291030"),
        ("OnePlus 12", 64999, "1938475620193840"),
        ("OnePlus 9 Pro", 64999, "5829103847562910"),
        ("OnePlus 11 5G", 56999, "2948571029384750"),
        ("OnePlus 12R", 39999, "9283746501928370"),
        ("OnePlus 11R", 35999, "1029384756102930"),
        ("OnePlus Nord 3", 33999, "4857291038475620"),
        ("OnePlus Nord 4", 29999, "7584930218475620"),
        ("OnePlus Nord CE 4", 24999, "3847562910384750"),
    ],
    "Apple": [
        ("iPhone 16 Pro Max", 159900, "8392048175930283"),
        ("iPhone 16 Pro", 134900, "4729104857291033"),
        ("iPhone 16", 90000, "1938475620193843"),
        ("iPhone 15 Pro Max", 139900, "5829103847562913"),
        ("iPhone 15 Pro", 114900, "2948571029384753"),
        ("iPhone 15 Plus", 89900, "3847562910384753"),
        ("iPhone 15", 79900, "9283746501928373"),
        ("iPhone 14 Pro Max", 129900, "1029384756102933"),
        ("iPhone 14 Pro", 99000, "4857291038475623"),
        ("iPhone 14", 60000, "7584930218475623"),
    ],
}

HEADERS = [
    "Brand", "Model", "Price", "Distributor_code",
    "Retailer_code", "Store_code", "EAN_code",
    "Currency", "Stock_units", "Sale_units"
]

countries = {
    "India":        { "prefix": "IND", "currency": "INR", "fx": 1.0,   "factor": 1.00, "gcs_folder": "INDIA" },
    "China":        { "prefix": "CHN", "currency": "CNY", "fx": 0.086, "factor": 1.02, "gcs_folder": "CHINA" },
    "America":      { "prefix": "USA", "currency": "USD", "fx": 0.012, "factor": 1.05, "gcs_folder": "AMERICA" },
    "South Korea":  { "prefix": "KOR", "currency": "KRW", "fx": 16.2,  "factor": 1.04, "gcs_folder": "SOUTH KOREA" },
    "Australia":    { "prefix": "AUS", "currency": "AUD", "fx": 0.018, "factor": 1.03, "gcs_folder": "AUSTRALIA" },
    "South Africa": { "prefix": "ZAF", "currency": "ZAR", "fx": 0.23,  "factor": 1.06, "gcs_folder": "SOUTH AFRICA" }
}

random.seed(42)

# COMMAND ----------
def local_price(price_inr, fx, factor):
    return max(1, int(round(price_inr * fx * factor)))

def rand_stock(price):
    if price >= 100000:
        s_min, s_max = 3, 20
    elif price >= 50000:
        s_min, s_max = 5, 35
    elif price >= 25000:
        s_min, s_max = 10, 60
    elif price >= 15000:
        s_min, s_max = 20, 90
    else:
        s_min, s_max = 40, 180
    
    stock = random.randint(s_min, s_max)
    sales = random.randint(1, stock)
    return stock, sales

# COMMAND ----------
# Start Generator Loop
num_distributors = 15
num_retailers = 10
brand_order = ["Samsung", "Vivo", "Oppo", "OnePlus", "Apple"]

for country_name, c in countries.items():
    logger.info(f"Generating data for Country: {country_name}")
    
    for d in range(1, num_distributors + 1):
        dist_code = f"{c['prefix']}DST{d:02d}"
        
        # Build retailer->stores hierarchy
        hierarchy = {}
        for r in range(1, num_retailers + 1):
            retailer_code = f"{dist_code}RET{r:02d}"
            hierarchy[retailer_code] = [
                f"{retailer_code}ST{s:02d}"
                for s in range(1, random.randint(5, 10) + 1)
            ]
            
        # Create Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for brand in brand_order:
            ws = wb.create_sheet(brand)
            
            # Write Header
            for col_idx, header in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
                
            row = 2
            for retailer_code, stores in hierarchy.items():
                for store_code in stores:
                    for model, price_inr, ean in CATALOGUE_INR[brand]:
                        price = local_price(price_inr, c["fx"], c["factor"])
                        stock_units, sale_units = rand_stock(price)
                        
                        values = [
                            brand, model, price, dist_code,
                            retailer_code, store_code, ean,
                            c["currency"], stock_units, sale_units
                        ]
                        
                        for col_idx, v in enumerate(values, 1):
                            ws.cell(row=row, column=col_idx, value=v)
                        row += 1
                        
        # Write directly to GCS memory buffer
        file_name = f"{dist_code}_{today_yyyymmdd}.xlsx"
        gcs_path = f"landing/{c['gcs_folder']}/{file_name}"
        
        # Save workbook to memory buffer and upload
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        blob = bucket.blob(gcs_path)
        blob.metadata = {"run_date": today_yyyymmdd, "country": c["gcs_folder"], "distributor": dist_code}
        blob.upload_from_string(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        logger.info(f"Uploaded: gs://{source_bucket_name}/{gcs_path}")

dbutils.notebook.exit("Successfully generated mock distributor files.")
